import sys
import datetime
import copy
import openpyxl
from openpyxl.styles import Border, Side

argument = sys.argv

print('using ' + argument[1])

wbname = argument[1]
# wbname = 'pythontest.xlsx'

wbnamesplited = wbname.split('.')

stBaseName = 'List'
stCreateSheetName = 'MasterSheet'
intUseColumnNum = 6
stUseColumnChoice = '使用'


dtCurrentDatetime = datetime.datetime.today()

#実行日を取得し、ファイル名として使用する
stExecuteTime = dtCurrentDatetime.strftime("%Y%m%d") 
stCreateFileName = wbnamesplited[0] + stExecuteTime + '.xlsx'

try:
    wb = openpyxl.load_workbook(wbname)
except:
    print('The file does not exist')

wb.create_sheet(stCreateSheetName)

wsBaseSheet = wb.get_sheet_by_name(stBaseName)
wsCreateSheet = wb.get_sheet_by_name(stCreateSheetName)

intBaseMaxRow = wsBaseSheet.max_row
intBaseMaxColumn = wsBaseSheet.max_column

#参照が渡されてしまって、Listシートが変更されてしまうのを修正したい
tbMaster = list(wsBaseSheet.rows)
tbBase = list(tbMaster)
del tbBase[0]

def fncCopyFirstColumns(ltBase,stColumn):

    stCheckValue = ""

    for c in ltBase:
        for r in c:
            if r.column == stColumn:
                if r.value == None:
                    r.value = stCheckValue
                stCheckValue = r.value
            else:
                break
        else:
            break

fncCopyFirstColumns(tbBase,"A")

def fncCopyNextColumns(ltBase,stColumn,stBaseColumn):

    stCheckValue = ""
    stFirstColumnValue = ""
        
    for c in ltBase:
        blCheck = False
        for r in c:
            if r.column == stColumn:
                if r.value == None and blCheck:
                    r.value = stCheckValue
                stCheckValue = r.value
            if r.column == stBaseColumn:
                if r.value == stFirstColumnValue:
                    blCheck = True
                stFirstColumnValue = r.value
            else:
                break              
        else:
            break

fncCopyNextColumns(tbBase,"B", "A")
fncCopyNextColumns(tbBase,"C", "B")
fncCopyNextColumns(tbBase,"D", "C")

ltUseList = []

if str(argument[2]) == "1":
    #使用する項目のみ取得
    for c in tbBase:
        intColumnCount = 0
        for r in c:
            intColumnCount += 1
            if intColumnCount == intUseColumnNum:
                if r.value == stUseColumnChoice:
                    ltUseList.append(c)
elif str(argument[2]) == "0":
    ltUseList = tbBase

intRowCount = 0

for row in ltUseList:
    intRowCount += 1
    intColumnCount = 0
    for cell in row:
        intColumnCount += 1
        if intColumnCount < intUseColumnNum:
            wsCreateSheet.cell(row=intColumnCount, column=intRowCount).value = cell.value
        else:
            break

tbUse = list(wsCreateSheet.rows)

border = Border(right=Side(border_style=medium,color='FF000000'))

for row in tbUse:
    stCheckValue = ""
    for cell in row:
        if stCheckValue != cell.value:
            cell.border = border
        stCheckValue = cell.value
        print(cell)
        cell.border = border

try:
    wb.save(stCreateFileName)
    print('The file was saved')
except:
    print('Erro!! The file can be opend')

